# -*- coding: utf-8 -*-
"""
Excel转JSON工具 - 为云开发数据库导入准备数据
"""
import openpyxl
import json
import os

# 列索引映射（来自data_loader.py）
COL = {
    '新增': 0,        '选科要求': 1,    '院校名称': 2,    '办学性质': 3,
    '专业名称': 4,    '计划数': 5,      '预测分数': 6,    '预测位次': 7,
    '24录取人数': 8,  '24最低分': 9,   '24最低位次': 10,
    '24学校最低分': 11,'24学校最低位次':12,'24学校平均分':13,'24学校平均位次':14,
    '23录取人数': 15, '23最低分': 16,  '23最低位次': 17,
    '22录取人数': 18, '22最低分': 19,  '22最低位次': 20,
    '院校水平_tag': 21,'院校水平': 22,  '省份': 23,      '城市': 24,
    '专业备注': 25,   '专业类': 26,     '学制': 27,      '学费': 28,
    '城市评级': 29,   '主管部门': 30,   '专业评估': 31,
    '硕士点': 32,     '硕士专业': 33,    '博士点': 34,     '博士专业': 35,
    '专业水平': 36,   '本专科': 37,     '院校排名': 38,
    '软科排名': 39,   '专业排名': 40,   '推免': 41,
    '保研率': 42,     '就业方向': 43,   '招生章程': 44,
    '学校招生信息': 45,'校园VR': 46,    '院校百科': 47,   '就业质量': 48,
}

EXCEL_PATH = r'C:\Users\DELL\.openclaw\workspace\data_raw.xlsx'
OUTPUT_DIR = r'C:\Users\DELL\.openclaw\workspace\cloudbase_data'

def get_row_value(row, col_name):
    """安全获取列值"""
    idx = COL.get(col_name)
    if idx is None:
        return None
    if idx >= len(row):
        return None
    val = row[idx]
    if isinstance(val, str):
        val = val.strip()
        if val == '':
            return None
    return val

def is_benke(row):
    """判断是否本科"""
    val = get_row_value(row, '本专科')
    return val == '本科'

def is_zhuanke(row):
    """判断是否专科"""
    val = get_row_value(row, '本专科')
    return val == '专科'

def convert_to_cloudbase_record(row, record_type):
    """转换为云数据库格式"""
    record = {
        "_id": f"{record_type}_{COL['院校名称']}_{COL['专业名称']}",  # 临时ID，导入后会生成
    }
    
    # 基础字段
    record['school_name'] = get_row_value(row, '院校名称')
    record['province'] = get_row_value(row, '省份')
    record['city'] = get_row_value(row, '城市')
    record['school_type'] = get_row_value(row, '办学性质')
    record['school_level'] = get_row_value(row, '院校水平')
    
    # 专业信息
    record['major_name'] = get_row_value(row, '专业名称')
    record['major_category'] = get_row_value(row, '专业类')
    record['study_system'] = get_row_value(row, '学制')
    record['tuition'] = get_row_value(row, '学费')
    record['major_note'] = get_row_value(row, '专业备注')
    
    # 选科要求
    record['subject_requirement'] = get_row_value(row, '选科要求')
    
    # 招生计划
    record['plan_2025'] = get_row_value(row, '计划数')
    
    # 预测分数/位次
    record['predicted_score_2025'] = get_row_value(row, '预测分数')
    record['predicted_rank_2025'] = get_row_value(row, '预测位次')
    
    # 2024录取数据
    record['adm2024_num'] = get_row_value(row, '24录取人数')
    record['adm2024_min_score'] = get_row_value(row, '24最低分')
    record['adm2024_min_rank'] = get_row_value(row, '24最低位次')
    record['adm2024_school_min_score'] = get_row_value(row, '24学校最低分')
    record['adm2024_school_min_rank'] = get_row_value(row, '24学校最低位次')
    record['adm2024_school_avg_score'] = get_row_value(row, '24学校平均分')
    record['adm2024_school_avg_rank'] = get_row_value(row, '24学校平均位次')
    
    # 2023录取数据
    record['adm2023_num'] = get_row_value(row, '23录取人数')
    record['adm2023_min_score'] = get_row_value(row, '23最低分')
    record['adm2023_min_rank'] = get_row_value(row, '23最低位次')
    
    # 2022录取数据
    record['adm2022_num'] = get_row_value(row, '22录取人数')
    record['adm2022_min_score'] = get_row_value(row, '22最低分')
    record['adm2022_min_rank'] = get_row_value(row, '22最低位次')
    
    # 院校排名
    record['school_rank'] = get_row_value(row, '院校排名')
    record['shuke_rank'] = get_row_value(row, '软科排名')
    record['major_rank'] = get_row_value(row, '专业排名')
    
    # 学术实力
    record['has_master'] = get_row_value(row, '硕士点')
    record['master_programs'] = get_row_value(row, '硕士专业')
    record['has_phd'] = get_row_value(row, '博士点')
    record['phd_programs'] = get_row_value(row, '博士专业')
    record['major_level'] = get_row_value(row, '专业水平')
    record['major_evaluation'] = get_row_value(row, '专业评估')
    
    # 就业信息
    record['employment_direction'] = get_row_value(row, '就业方向')
    record['employment_quality'] = get_row_value(row, '就业质量')
    
    # 招生章程
    record['admission_rules'] = get_row_value(row, '招生章程')
    record['school_admission_info'] = get_row_value(row, '学校招生信息')
    record['school_vr'] = get_row_value(row, '校园VR')
    record['school_wiki'] = get_row_value(row, '院校百科')
    
    # 城市评级
    record['city_level'] = get_row_value(row, '城市评级')
    
    # 主管部门
    record['governing_body'] = get_row_value(row, '主管部门')
    
    # 保研率
    record[' exemption_rate'] = get_row_value(row, '保研率')
    
    # 删除None值
    record = {k: v for k, v in record.items() if v is not None}
    
    return record

def main():
    # 创建输出目录
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    print(f"Loading: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    
    benke_records = []
    zhuanke_records = []
    
    print(f"Total rows: {ws.max_row - 1}")
    
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i % 2000 == 0:
            print(f"Processing row {i}...")
        
        if is_benke(row):
            record = convert_to_cloudbase_record(row, 'bk')
            benke_records.append(record)
        elif is_zhuanke(row):
            record = convert_to_cloudbase_record(row, 'zk')
            zhuanke_records.append(record)
    
    wb.close()
    
    # 保存JSON文件（支持云开发导入格式）
    benke_path = os.path.join(OUTPUT_DIR, 'benke.json')
    zhuanke_path = os.path.join(OUTPUT_DIR, 'zhuanke.json')
    
    with open(benke_path, 'w', encoding='utf-8') as f:
        json.dump(benke_records, f, ensure_ascii=False, indent=2)
    
    with open(zhuanke_path, 'w', encoding='utf-8') as f:
        json.dump(zhuanke_records, f, ensure_ascii=False, indent=2)
    
    print(f"\n=== 转换完成 ===")
    print(f"本科: {len(benke_records)} 条记录 -> {benke_path}")
    print(f"专科: {len(zhuanke_records)} 条记录 -> {zhuanke_path}")
    
    # 生成导入说明
    guide = f"""
=== 云开发数据导入指南 ===

## 文件位置
- 本科数据: {benke_path}
- 专科数据: {zhuanke_path}

## 导入步骤

### 1. 导入本科数据
1. 打开微信开发者工具
2. 点击"数据库" -> "benke"集合
3. 点击"导入"按钮
4. 选择 benke.json 文件
5. 导入格式选择: JSON
6. 点击确认

### 2. 导入专科数据
同上，选择 zhuanke.json 导入到 zhuanke 集合

## 数据字段说明
- school_name: 院校名称
- major_name: 专业名称
- province: 省份
- city: 城市
- school_type: 办学性质（公办/民办）
- predicted_score_2025: 2025预测分数
- predicted_rank_2025: 2025预测位次
- adm2024_min_score: 2024最低分
- adm2023_min_score: 2023最低分
- adm2022_min_score: 2022最低分
...等共49个字段

## 注意
- 导入前请确保集合存在
- JSON格式为数组格式
- 每条记录会自动生成_id
"""
    
    guide_path = os.path.join(OUTPUT_DIR, 'IMPORT_GUIDE.txt')
    with open(guide_path, 'w', encoding='utf-8') as f:
        f.write(guide)
    
    print(f"导入指南: {guide_path}")

if __name__ == '__main__':
    main()
